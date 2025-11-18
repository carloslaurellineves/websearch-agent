"""Escritor de arquivos Excel com formatação."""

import logging
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from src.models.software import SoftwareResult

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Classe para escrever resultados em arquivos Excel com formatação."""

    def __init__(self, output_path: Path):
        """
        Inicializa o escritor de Excel.

        Args:
            output_path: Caminho do arquivo de saída
        """
        self.output_path = output_path

    def write_results(self, results: List[SoftwareResult]) -> bool:
        """
        Escreve os resultados em um arquivo Excel com formatação.

        Args:
            results: Lista de SoftwareResult

        Returns:
            True se escrita bem-sucedida, False caso contrário
        """
        try:
            logger.info(f"Escrevendo {len(results)} resultados em: {self.output_path}")

            # Garante que o diretório existe
            self.output_path.parent.mkdir(parents=True, exist_ok=True)

            # Converte resultados para DataFrame
            data = [result.to_excel_row() for result in results]
            df = pd.DataFrame(data)

            # Define a ordem das colunas
            columns = [
                "Nome",
                "Versão",
                "Status Original",
                "Status Verificado",
                "Data Pesquisa",
                "Fontes",
                "Links",
                "Confiança",
                "Resumo",
            ]
            df = df[columns]

            # Salva o Excel
            df.to_excel(self.output_path, index=False, engine="openpyxl")

            # Aplica formatação
            self._apply_formatting(self.output_path)

            logger.info(f"Arquivo Excel criado com sucesso: {self.output_path}")
            return True

        except Exception as e:
            logger.error(f"Erro ao escrever arquivo Excel: {e}")
            return False

    def _apply_formatting(self, file_path: Path) -> None:
        """
        Aplica formatação ao arquivo Excel.

        Args:
            file_path: Caminho do arquivo Excel
        """
        try:
            wb = load_workbook(file_path)
            ws = wb.active

            # Formatação do cabeçalho
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Ajusta largura das colunas
            column_widths = {
                "A": 30,  # Nome
                "B": 15,  # Versão
                "C": 18,  # Status Original
                "D": 18,  # Status Verificado
                "E": 20,  # Data Pesquisa
                "F": 25,  # Fontes
                "G": 40,  # Links
                "H": 12,  # Confiança
                "I": 50,  # Resumo
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # Formatação condicional baseada em confiança e status
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                # Coluna de Confiança (H)
                confianca_cell = ws[f"H{row_idx}"]
                confianca_value = confianca_cell.value

                if isinstance(confianca_value, (int, float)):
                    if confianca_value >= 80:
                        confianca_cell.fill = PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                        )
                    elif confianca_value >= 50:
                        confianca_cell.fill = PatternFill(
                            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                        )
                    else:
                        confianca_cell.fill = PatternFill(
                            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                        )

                # Coluna de Status Verificado (D)
                status_cell = ws[f"D{row_idx}"]
                status_value = str(status_cell.value or "").upper()

                if status_value == "SIM":
                    status_cell.fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )
                    status_cell.font = Font(bold=True, color="9C0006")
                elif status_value == "NÃO" or status_value == "NAO":
                    status_cell.fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )
                    status_cell.font = Font(bold=True, color="006100")
                elif status_value == "ERRO":
                    status_cell.fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )
                    status_cell.font = Font(bold=True, color="000000")

                # Alinhamento de células
                for cell in row:
                    if cell.column_letter in ["H"]:  # Confiança - centralizado
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif cell.column_letter in ["D", "C"]:  # Status - centralizado
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:  # Outras colunas - alinhamento à esquerda
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # Congela a primeira linha
            ws.freeze_panes = "A2"

            # Salva as alterações
            wb.save(file_path)
            logger.debug("Formatação aplicada com sucesso")

        except Exception as e:
            logger.warning(f"Erro ao aplicar formatação (arquivo ainda foi criado): {e}")

